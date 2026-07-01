"""Korean public-data local-agent file-discovery benchmark builder.

The builder creates a bounded, reproducible local corpus from public open-data
spreadsheet downloads and turns it into a messy-folder search benchmark for
local agents.  The default downloader uses Seoul Data Hub because it exposes
direct public XLSX downloads without API credentials; the manifest records that
the run is a Public Data Portal-style fallback when data.go.kr itself is not
reachable from the machine.
"""
from __future__ import annotations

import random
import re
import shutil
import time
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass
from html import unescape
from io import BytesIO
from pathlib import Path
from typing import Any

from .agent_index import build_agent_index
from .config import Config
from .eval import _write_json, _write_jsonl
from .hippocamp import BenchResult, run_benchmark

SEOUL_BASE = "https://data.seoul.go.kr/bsp"
SEOUL_VIEW = SEOUL_BASE + "/wgs/dataView/data300View/{id}.do"
SEOUL_XLSX = SEOUL_BASE + "/wgs/dataset/dataXlsxDown.do"
DEFAULT_PUBLICDATA_SEED = 20260529
PUBLICDATA_QUERY_STOP_TERMS = {
    "sheet",
    "sheet1",
    "data",
    "null",
    "none",
    "서울",
    "서울시",
    "통계",
    "현황",
    "데이터",
    "공공데이터",
    "연도",
    "구별",
    "동별",
    "성별",
    "합계",
    "총계",
}


@dataclass
class PublicDataBuildResult:
    dest: Path
    train_root: Path
    valid_root: Path
    test_root: Path
    train_eval_set_path: Path
    valid_eval_set_path: Path
    eval_set_path: Path
    manifest_path: Path
    docs_downloaded: int
    train_docs: int
    valid_docs: int
    test_docs: int
    eval_cases: int


@dataclass
class PublicDataSuiteResult:
    build: PublicDataBuildResult
    deterministic_report: Path
    deterministic_metrics: dict[str, Any]
    prepare_seconds: float
    report_path: Path


def _slug(value: str, *, max_len: int = 80) -> str:
    value = re.sub(r"[\\/:*?\"<>|]+", " ", value or "")
    value = re.sub(r"\s+", "_", value).strip("._ ")
    return (value or "untitled")[:max_len]


def _http_get(url: str, *, timeout: int = 45) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 Jikji-publicdata-bench"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:  # noqa: S310 - public benchmark URL.
        return resp.read()


def _http_post(url: str, data: dict[str, str], *, timeout: int = 90) -> bytes:
    payload = urllib.parse.urlencode(data).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            "User-Agent": "Mozilla/5.0 Jikji-publicdata-bench",
            "Referer": SEOUL_VIEW.format(id=data.get("id", "")),
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:  # noqa: S310 - public benchmark URL.
        return resp.read()


def _text_between(html: str, pattern: str) -> str:
    match = re.search(pattern, html, flags=re.DOTALL)
    if not match:
        return ""
    raw = re.sub(r"<[^>]+>", " ", match.group(1))
    return re.sub(r"\s+", " ", unescape(raw)).strip()


def _metadata_from_html(dataset_id: int, html: str) -> dict[str, str]:
    title = _text_between(html, r'<h1 class="trgtTableIndctNm">(.+?)</h1>')
    desc = _text_between(html, r'<div class="dsDesc">(.+?)</div>')
    category = _text_between(html, r'<div class="type-title[^"]* lvl1JrmCd[^"]*">(.+?)</div>')
    return {
        "source": "Seoul Data Hub public XLSX download",
        "source_url": SEOUL_VIEW.format(id=dataset_id),
        "download_url": SEOUL_XLSX,
        "dataset_id": str(dataset_id),
        "title": title or f"dataset-{dataset_id}",
        "description": desc,
        "category": category,
        "license_note": (
            "Public open-data download. Requested source was data.go.kr KOGL Type 1; "
            "this run records Seoul Data Hub as an accessible public-data fallback."
        ),
    }


def _looks_like_xlsx(blob: bytes) -> bool:
    if len(blob) < 2048 or not blob.startswith(b"PK"):
        return False
    try:
        with zipfile.ZipFile(BytesIO(blob)) as zf:
            return "[Content_Types].xml" in zf.namelist()
    except (OSError, zipfile.BadZipFile):
        return False


def _xlsx_text(path: Path, *, max_cells: int = 240) -> list[str]:
    """Extract enough visible text from XLSX for benchmark query generation."""
    out: list[str] = []
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames[:6]:
                out.append(sheet_name)
                ws = wb[sheet_name]
                for row in ws.iter_rows(max_row=50, values_only=True):
                    for cell in row:
                        if cell is None:
                            continue
                        text = str(cell).strip()
                        if text:
                            out.append(text)
                        if len(out) >= max_cells:
                            return out
        finally:
            wb.close()
    except Exception:
        pass
    return out


def _tokens(text: str, *, min_len: int = 2) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for match in re.finditer(r"[0-9A-Za-z가-힣][0-9A-Za-z가-힣_.+-]*", text or ""):
        token = match.group(0).strip("._+-")
        if len(token) < min_len or token.casefold() in PUBLICDATA_QUERY_STOP_TERMS or token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def _rare_terms(docs: list[dict[str, Any]], *, limit_per_doc: int = 6) -> dict[str, list[str]]:
    freq: dict[str, int] = {}
    doc_terms: dict[str, list[str]] = {}
    for doc in docs:
        terms = _tokens(" ".join(doc.get("xlsx_text") or []), min_len=3)
        doc_terms[doc["bench_path"]] = terms
        for term in set(terms):
            freq[term] = freq.get(term, 0) + 1
    out: dict[str, list[str]] = {}
    for path, terms in doc_terms.items():
        ranked = sorted(set(terms), key=lambda t: (freq.get(t, 9999), -len(t), t))
        out[path] = [
            term for term in ranked
            if freq.get(term, 9999) <= 2 and term.casefold() not in PUBLICDATA_QUERY_STOP_TERMS
        ][:limit_per_doc]
    return out


def _messy_relpath(doc: dict[str, Any], split: str, idx: int, rng: random.Random) -> str:
    buckets = [
        ("01_업무공유", "정리전", "엑셀 원본"),
        ("받은자료", "기관별", "새 폴더"),
        ("데이터_검토", "2024_하반기", "원본"),
        ("공공데이터", "임시보관", "확인필요"),
        ("team_drive", "공유받음", "misc"),
    ]
    first, second, third = buckets[idx % len(buckets)]
    title = _slug(str(doc.get("title") or f"dataset-{idx}"), max_len=48)
    suffix = rng.choice(["최종", "수정본", "원본", "검토용", "다운로드"])
    version = rng.choice(["v1", "v2", "2024", "2025", "last"])
    name = f"{idx:03d}_{title}_{suffix}_{version}.xlsx"
    return f"{split}/{first}/{second}/{third}/{name}"


def _download_publicdata_docs(dest: Path, *, target_docs: int, max_id: int, seed: int) -> list[dict[str, Any]]:
    source_dir = dest / "source_downloads"
    source_dir.mkdir(parents=True, exist_ok=True)
    rng = random.Random(seed)
    ids = list(range(1, max_id + 1))
    # Keep low IDs first because Seoul's public spreadsheet endpoint is denser
    # there, but jitter within blocks to avoid one topical cluster only.
    ids = [i for block in range(0, len(ids), 50) for i in rng.sample(ids[block:block + 50], len(ids[block:block + 50]))]
    docs: list[dict[str, Any]] = []
    failures: list[dict[str, str]] = []
    for dataset_id in ids:
        if len(docs) >= target_docs:
            break
        xlsx_path = source_dir / f"seoul_{dataset_id}.xlsx"
        html_path = source_dir / f"seoul_{dataset_id}.html"
        try:
            if xlsx_path.exists() and xlsx_path.stat().st_size > 2048:
                blob = xlsx_path.read_bytes()
            else:
                blob = _http_post(
                    SEOUL_XLSX,
                    {"id": str(dataset_id), "tdColNmArr": "", "rowFilterList": "[]"},
                )
                if not _looks_like_xlsx(blob):
                    failures.append({"id": str(dataset_id), "reason": "not_xlsx"})
                    continue
                xlsx_path.write_bytes(blob)
            try:
                html = html_path.read_text(encoding="utf-8")
            except OSError:
                html = _http_get(SEOUL_VIEW.format(id=dataset_id)).decode("utf-8", "ignore")
                html_path.write_text(html, encoding="utf-8")
            meta = _metadata_from_html(dataset_id, html)
            meta["source_file"] = str(xlsx_path)
            meta["bytes"] = str(len(blob))
            meta["xlsx_text"] = _xlsx_text(xlsx_path)
            docs.append(meta)
        except Exception as exc:  # network sites can fail per-id; keep bounded crawl going.
            failures.append({"id": str(dataset_id), "reason": type(exc).__name__})
            continue
    _write_json(dest / "download_failures.json", {"failures": failures[:200], "count": len(failures)})
    return docs


def _split_docs(docs: list[dict[str, Any]], *, seed: int) -> dict[str, list[dict[str, Any]]]:
    shuffled = list(docs)
    random.Random(seed).shuffle(shuffled)
    n = len(shuffled)
    train_end = max(1, int(n * 0.6))
    valid_end = max(train_end + 1, int(n * 0.8))
    return {
        "train": shuffled[:train_end],
        "valid": shuffled[train_end:valid_end],
        "test": shuffled[valid_end:],
    }


def _materialize_split(dest: Path, split: str, docs: list[dict[str, Any]], *, seed: int) -> list[dict[str, Any]]:
    rng = random.Random(seed + len(split))
    root = dest / "corpus" / split
    root.mkdir(parents=True, exist_ok=True)
    materialized: list[dict[str, Any]] = []
    for idx, doc in enumerate(docs, 1):
        rel = _messy_relpath(doc, split, idx, rng)
        target = root / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(Path(str(doc["source_file"])), target)
        row = dict(doc)
        row["split"] = split
        row["bench_path"] = rel
        row["materialized_path"] = str(target)
        materialized.append(row)
        # Human-ish clutter: adjacent note files with similar words but not the
        # expected spreadsheet.  This makes raw folder browsing noisier.
        if idx % 4 == 0:
            note = target.with_suffix(".메모.txt")
            note.write_text(
                f"임시 메모: {doc.get('title')} 자료 검토 중. 원본 엑셀은 같은 폴더에 있음.\n",
                encoding="utf-8",
            )
    return materialized


def _case_templates(test_docs: list[dict[str, Any]], *, max_cases: int) -> list[dict[str, Any]]:
    rare_by_path = _rare_terms(test_docs)
    cases: list[dict[str, Any]] = []
    scenario_cycle = [
        "filename_vague",
        "content_lexical",
        "semantic_description",
        "folder_context",
        "column_or_value",
    ]
    for idx, doc in enumerate(test_docs):
        if len(cases) >= max_cases:
            break
        path = str(doc["bench_path"])
        title = str(doc.get("title") or Path(path).stem)
        desc = str(doc.get("description") or "")
        folder = "/".join(Path(path).parts[:4])
        rare = rare_by_path.get(path) or _tokens(" ".join(doc.get("xlsx_text") or []), min_len=3)[:3]
        scenario = scenario_cycle[idx % len(scenario_cycle)]
        if scenario == "filename_vague":
            query = f"전에 받은 공공데이터 엑셀 중 제목이 '{title[:35]}' 비슷했던 원본 파일 찾아줘"
        elif scenario == "content_lexical":
            clue = rare[0] if rare else title.split()[0]
            query = f"엑셀 본문 안에 '{clue}' 값이나 항목이 들어간 공공데이터 파일 찾아줘"
        elif scenario == "semantic_description":
            terms = " ".join(_tokens(f"{title} {desc}", min_len=2)[:8])
            query = f"파일명은 정확히 모르지만 {terms} 관련 현황을 담은 데이터셋을 찾아줘"
        elif scenario == "folder_context":
            query = f"{folder} 쪽에 정리해 둔 {title[:30]} 관련 엑셀 원본을 찾아줘"
        else:
            clue = rare[1] if len(rare) > 1 else (rare[0] if rare else title[:12])
            query = f"컬럼이나 행 값으로 '{clue}' 단서가 보이는 스프레드시트를 찾아줘"
        cases.append({
            "id": f"publicdata-{idx + 1:04d}",
            "scenario": scenario,
            "query": query,
            "expected_paths": [path],
            "expected_source_url": doc.get("source_url", ""),
            "dataset_title": title,
            "license_note": doc.get("license_note", ""),
            "public_benchmark": True,
        })
    return cases


def build_publicdata_benchmark(
    dest: Path,
    *,
    target_docs: int = 90,
    max_id: int = 700,
    max_cases: int = 40,
    seed: int = DEFAULT_PUBLICDATA_SEED,
) -> PublicDataBuildResult:
    dest = Path(dest).expanduser().resolve()
    dest.mkdir(parents=True, exist_ok=True)
    docs = _download_publicdata_docs(dest, target_docs=target_docs, max_id=max_id, seed=seed)
    if len(docs) < 12:
        raise RuntimeError(f"Too few public-data documents downloaded: {len(docs)}")
    splits = _split_docs(docs, seed=seed)
    materialized: dict[str, list[dict[str, Any]]] = {}
    for split, rows in splits.items():
        materialized[split] = _materialize_split(dest, split, rows, seed=seed)
        _write_jsonl(dest / "metadata" / f"{split}_docs.jsonl", materialized[split])
    eval_sets: dict[str, Path] = {}
    eval_counts: dict[str, int] = {}
    for split, rows in materialized.items():
        split_cases = _case_templates(rows, max_cases=max_cases)
        split_eval_set = dest / "eval" / f"publicdata_{split}_eval.jsonl"
        _write_jsonl(split_eval_set, split_cases)
        eval_sets[split] = split_eval_set
        eval_counts[split] = len(split_cases)
    eval_set = eval_sets["test"]
    manifest = dest / "manifest.json"
    _write_json(manifest, {
        "source_family": "Korean public open data",
        "requested_source": "data.go.kr Public Data Portal / KOGL Type 1 documents",
        "actual_source": "Seoul Data Hub public XLSX endpoint",
        "honesty_note": (
            "data.go.kr direct download was not assumed because it may require credentials/session access; "
            "this builder records Seoul Data Hub public-data XLSX files as an accessible fallback."
        ),
        "seed": seed,
        "docs_downloaded": len(docs),
        "splits": {split: len(rows) for split, rows in materialized.items()},
        "eval_set": str(eval_set),
        "eval_sets": {split: str(path) for split, path in eval_sets.items()},
        "eval_cases": eval_counts["test"],
        "eval_case_counts": eval_counts,
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    })
    return PublicDataBuildResult(
        dest=dest,
        train_root=dest / "corpus" / "train",
        valid_root=dest / "corpus" / "valid",
        test_root=dest / "corpus" / "test",
        train_eval_set_path=eval_sets["train"],
        valid_eval_set_path=eval_sets["valid"],
        eval_set_path=eval_set,
        manifest_path=manifest,
        docs_downloaded=len(docs),
        train_docs=len(materialized["train"]),
        valid_docs=len(materialized["valid"]),
        test_docs=len(materialized["test"]),
        eval_cases=eval_counts["test"],
    )


def run_publicdata_suite(
    dest: Path,
    *,
    target_docs: int = 90,
    max_id: int = 700,
    max_cases: int = 40,
    seed: int = DEFAULT_PUBLICDATA_SEED,
    top_k: int = 10,
) -> PublicDataSuiteResult:
    build = build_publicdata_benchmark(
        dest,
        target_docs=target_docs,
        max_id=max_id,
        max_cases=max_cases,
        seed=seed,
    )
    cfg = Config(include_hidden=False)
    cfg.max_files = 1_000_000
    t0 = time.perf_counter()
    build_agent_index(build.train_root, cfg)
    build_agent_index(build.valid_root, cfg)
    build_agent_index(build.test_root, cfg)
    prepare_seconds = time.perf_counter() - t0
    bench: BenchResult = run_benchmark(
        build.test_root,
        eval_set=build.eval_set_path,
        modes=("raw", "jikji"),
        top_k=top_k,
        prepare=False,
        allow_leak=False,
    )
    report_path = Path(dest).expanduser().resolve() / "reports" / "publicdata_suite_report.json"
    _write_json(report_path, {
        "build": {
            "dest": str(build.dest),
            "train_root": str(build.train_root),
            "valid_root": str(build.valid_root),
            "test_root": str(build.test_root),
            "train_eval_set": str(build.train_eval_set_path),
            "valid_eval_set": str(build.valid_eval_set_path),
            "eval_set": str(build.eval_set_path),
            "docs_downloaded": build.docs_downloaded,
            "train_docs": build.train_docs,
            "valid_docs": build.valid_docs,
            "test_docs": build.test_docs,
            "eval_cases": build.eval_cases,
        },
        "prepare_seconds": round(prepare_seconds, 3),
        "deterministic_report": str(bench.report_path),
        "deterministic_metrics": bench.metrics,
    })
    return PublicDataSuiteResult(
        build=build,
        deterministic_report=bench.report_path,
        deterministic_metrics=bench.metrics,
        prepare_seconds=round(prepare_seconds, 3),
        report_path=report_path,
    )
